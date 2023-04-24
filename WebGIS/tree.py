import os
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import re
import openpyxl


root = tk.Tk()
root.withdraw()

messagebox.showinfo('提示','请选择Excel！')
datapath = filedialog.askopenfilename() 

wb = openpyxl.load_workbook(datapath)

sh = wb[wb.sheetnames[0]]
rows = sh.max_row
for i in range(1,rows+1):
    try:
        if sh.cell(i,1).value == 0 :
            sh.cell(i,1).value = 'Z00000000'
        else :
            old_str = str(re.sub(r'[0-9]+', '', sh.cell(i,1).value))
            new_str = ''
            for j in range(len(old_str)) :
                if old_str[j] == 's' or old_str[j] == 'S' :
                    new_str = new_str.strip('H') + 'S'
                else :
                    if old_str[j] == 'j' or old_str[j] == 'J' :
                        new_str = new_str + old_str[j]
                    elif old_str[j] == 'x' or old_str[j] == 'X' :
                        new_str = new_str + old_str[j]
                    else :
                        new_str = new_str + old_str[j] + 'H'

            if new_str[0] == 'j' or new_str[0] == 'J' :
                sh.cell(i,1).value  = new_str
            elif new_str[0] == 'x' or new_str[0] == 'X' :
                sh.cell(i,1).value  = new_str
            else :
                sh.cell(i,1).value = 'Z'+new_str
    except:
        print('----------------',sh.cell(i,1).value)
messagebox.showinfo('提示','填充完成！')
wb.save(datapath)
    