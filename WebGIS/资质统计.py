import os
from tkinter import filedialog
from docx import Document
import tkinter as tk
from tkinter import messagebox
import pandas as pd
from openpyxl import Workbook

root = tk.Tk()
root.withdraw()

messagebox.showinfo('提示','请选择文件夹！')

Folderpath = filedialog.askdirectory() #获得选择好的文件夹

# 获取filepath文件夹下的所有的文件
def getfilelist(filepath):
    filelist =  os.listdir(filepath)
    files = []
    for i in range(len(filelist)):
        child = os.path.join('%s/%s'%(filepath, filelist[i]))
        if os.path.isdir(child):
            files.extend(getfilelist(child))
        else:
            files.append(child)
    return files

lists = []
for filepath in getfilelist(Folderpath):
    if filepath.endswith("docx") and not filepath.startswith('~$'):
        lists.append(filepath)


workbook = Workbook()
worksheet = workbook.active
worksheet.cell(1,1).value = '单位名称'
worksheet.cell(1,2).value = '申请资质等级类别'
i = 1
for filepath in lists:
    try:
        i+=1
        doc = Document(filepath)
        table = doc.tables[0]
        s1 = table.cell(0,1).text
        s2 = table.cell(5,1).text
        worksheet.cell(i,1).value = s1
        worksheet.cell(i,2).value = s2
        if (i== 39) :
            
            print('-====',filepath )
    except:
        print('问题路径---',filepath )
        continue
workbook.save(Folderpath + '/' + '资质统计.xlsx')
messagebox.showinfo('提示','资质等级统计完成！')