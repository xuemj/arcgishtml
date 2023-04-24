# -*- coding: utf-8 -*-
import os
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Border, Side, Alignment, Font
alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)

root = tk.Tk()
root.withdraw()

messagebox.showinfo('提示','请选择需要整合的Excel表！')
datapath = filedialog.askopenfilename() 
#读取文档
df1 = pd.read_excel(datapath, sheet_name=0)
xzqdm_list = df1['XZQDM'].tolist()
xzqmc_list = df1['XZQMC'].tolist()
dlmc_list = df1['DLMC'].tolist()
zldwdm_list = df1['ZLDWDM'].tolist()
pjjg_list = df1['PJJG'].tolist()
# dlbm_list = df1['DLBM'].tolist()
bz_list = df1['BZ'].tolist()
zldwmc_list = df1['ZLDWMC'].tolist()
tbdlmj_list = df1['TBDLMJ'].tolist()

# 创建一个workbook 
workbook = Workbook()
# 创建一个worksheet
worksheet = workbook.active
file_excel = os.path.dirname(datapath)
worksheet.column_dimensions['A'].width = 22
worksheet.column_dimensions['B'].width = 19
worksheet.column_dimensions['C'].width = 18
worksheet.column_dimensions['D'].width = 13.38
worksheet.column_dimensions['E'].width = 13.38
worksheet.column_dimensions['F'].width = 13.38
worksheet.column_dimensions['G'].width = 13.38
worksheet.column_dimensions['H'].width = 13.38
worksheet.column_dimensions['I'].width = 13.38
worksheet.column_dimensions['J'].width = 13.38
worksheet.column_dimensions['K'].width = 13.38
worksheet.column_dimensions['L'].width = 13.38
worksheet.column_dimensions['M'].width = 13.38
worksheet.column_dimensions['N'].width = 13.38

worksheet.merge_cells('A1:N1')
worksheet.cell(1,1).value = '济源市耕地后备资源调查评价分类型统计表'
worksheet.cell(1,1).alignment = alignment
worksheet.merge_cells('A2:N2')
worksheet.cell(2,1).value = '单位：公顷'
worksheet.cell(2,1).alignment = Alignment(horizontal='right',vertical='center')
worksheet.cell(3, 1).value = '行政区   代码'
worksheet.cell(3, 1).alignment = alignment
worksheet.cell(3, 2).value = '行政区   名称'
worksheet.cell(3, 2).alignment = alignment
worksheet.cell(3, 3).value = '合计'
worksheet.cell(3, 3).alignment = alignment
worksheet.cell(3, 4).value = '宜耕小计'
worksheet.cell(3, 4).alignment = alignment
worksheet.cell(3, 5).value = '宜耕\n其他草地'
worksheet.cell(3, 5).alignment = alignment
worksheet.cell(3, 6).value = '宜耕\n盐碱地'
worksheet.cell(3, 6).alignment = alignment
worksheet.cell(3, 7).value = '宜耕\n沙地'
worksheet.cell(3, 7).alignment = alignment
worksheet.cell(3, 8).value = '宜耕\n裸地'
worksheet.cell(3, 8).alignment = alignment
worksheet.cell(3, 9).value = '不宜耕小计'
worksheet.cell(3, 9).alignment = alignment
worksheet.cell(3, 10).value = '不宜耕\n其他草地'
worksheet.cell(3, 10).alignment = alignment
worksheet.cell(3, 11).value = '不宜耕\n盐碱地'
worksheet.cell(3, 11).alignment = alignment
worksheet.cell(3, 12).value = '不宜耕\n沙地'
worksheet.cell(3, 12).alignment = alignment
worksheet.cell(3, 13).value = '不宜耕\n裸地'
worksheet.cell(3, 13).alignment = alignment
worksheet.cell(3, 14).value = '其他'
worksheet.cell(3, 14).alignment = alignment

index  = 4
row = 0
xzqdm = 0
zldwdm = 0
total_index = 0
total_arr = []
for i in range(len(xzqdm_list)) :
    row = row + 1
    if xzqdm_list[i] != xzqdm :
        if total_index > 0 :
            t3=t4=t5=t6=t7=t8=t9=t10=t11=t12=t13=t14=0
            for p in range(total_index+1,row+index) :
                t3 = t3 + worksheet.cell(p,3).value
                t4 = t4 + worksheet.cell(p,4).value
                t5 = t5 + worksheet.cell(p,5).value
                t6 = t6 + worksheet.cell(p,6).value
                t7 = t7 + worksheet.cell(p,7).value
                t8 = t8 + worksheet.cell(p,8).value
                t9 = t9 + worksheet.cell(p,9).value
                t10 = t10 + worksheet.cell(p,10).value
                t11 = t11 + worksheet.cell(p,11).value
                t12 = t12 + worksheet.cell(p,12).value
                t13 = t13 + worksheet.cell(p,13).value
                t14 = t14 + worksheet.cell(p,14).value
            worksheet.cell(total_index,3).value = t3
            worksheet.cell(total_index,4).value = t4
            worksheet.cell(total_index,5).value = t5
            worksheet.cell(total_index,6).value = t6
            worksheet.cell(total_index,7).value = t7
            worksheet.cell(total_index,8).value = t8
            worksheet.cell(total_index,9).value = t9
            worksheet.cell(total_index,10).value = t10
            worksheet.cell(total_index,11).value = t11
            worksheet.cell(total_index,12).value = t12
            worksheet.cell(total_index,13).value = t13
            worksheet.cell(total_index,14).value = t14
            
        worksheet.cell(row+index,1).value = str(xzqdm_list[i])
        worksheet.cell(row+index,2).value = str(xzqmc_list[i])
        worksheet.cell(row+index,3).value = 0.0000
        worksheet.cell(row+index,4).value = 0.0000
        worksheet.cell(row+index,5).value = 0.0000
        worksheet.cell(row+index,6).value = 0.0000
        worksheet.cell(row+index,7).value = 0.0000
        worksheet.cell(row+index,8).value = 0.0000
        worksheet.cell(row+index,9).value = 0.0000
        worksheet.cell(row+index,10).value = 0.0000 
        worksheet.cell(row+index,11).value = 0.0000
        worksheet.cell(row+index,12).value = 0.0000
        worksheet.cell(row+index,13).value = 0.0000
        worksheet.cell(row+index,14).value = 0.0000
        total_index = row+index
        total_arr.append(total_index)
        xzqdm = xzqdm_list[i]
        row = row + 1
    if zldwdm_list[i] != zldwdm :
        worksheet.cell(row+index,1).value = str(zldwdm_list[i])
        worksheet.cell(row+index,2).value = str(zldwmc_list[i])
        worksheet.cell(row+index,3).value = 0.0000
        worksheet.cell(row+index,4).value = 0.0000
        worksheet.cell(row+index,5).value = 0.0000
        worksheet.cell(row+index,6).value = 0.0000
        worksheet.cell(row+index,7).value = 0.0000
        worksheet.cell(row+index,8).value = 0.0000
        worksheet.cell(row+index,9).value = 0.0000
        worksheet.cell(row+index,10).value = 0.0000
        worksheet.cell(row+index,11).value = 0.0000
        worksheet.cell(row+index,12).value = 0.0000
        worksheet.cell(row+index,13).value = 0.0000
        worksheet.cell(row+index,14).value = 0.0000
        if dlmc_list[i] == '其他草地' and pjjg_list[i] != ' ':
            if pjjg_list[i] == 'FHB':
                worksheet.cell(row+index,10).value = tbdlmj_list[i]
            else :
                worksheet.cell(row+index,5).value = tbdlmj_list[i]
        elif dlmc_list[i] == '盐碱地' and pjjg_list[i] != ' ':
            if pjjg_list[i] == 'FHB' :
                worksheet.cell(row+index,11).value = tbdlmj_list[i]
            else :
                worksheet.cell(row+index,6).value = tbdlmj_list[i]
        elif dlmc_list[i] == '沙地' and pjjg_list[i] != ' ':
            if pjjg_list[i] == 'FHB' :
                worksheet.cell(row+index,12).value = tbdlmj_list[i]
            else :
                worksheet.cell(row+index,7).value = tbdlmj_list[i]
        elif dlmc_list[i] == '裸土地' and pjjg_list[i] != ' ':
            if pjjg_list[i] == 'FHB' :
                worksheet.cell(row+index,13).value = tbdlmj_list[i]
            else :
                worksheet.cell(row+index,8).value = tbdlmj_list[i]
        else :
            worksheet.cell(row+index,14).value = tbdlmj_list[i]
        # 相关合计
        worksheet.cell(row+index,4).value = worksheet.cell(row+index,5).value +worksheet.cell(row+index,6).value+worksheet.cell(row+index,7).value+worksheet.cell(row+index,8).value
        worksheet.cell(row+index,9).value = worksheet.cell(row+index,10).value +worksheet.cell(row+index,11).value+worksheet.cell(row+index,12).value+worksheet.cell(row+index,13).value
        worksheet.cell(row+index,3).value = worksheet.cell(row+index,4).value+worksheet.cell(row+index,9).value+worksheet.cell(row+index,14).value
        zldwdm = zldwdm_list[i]
    else :
        row = row -1
        if dlmc_list[i] == '其他草地' and pjjg_list[i] != ' ':
            if pjjg_list[i] == 'FHB' :
                if worksheet.cell(row+index,10).value != 0.0000:
                    worksheet.cell(row+index,10).value = tbdlmj_list[i] + float(worksheet.cell(row+index,10).value)
                else :
                    worksheet.cell(row+index,10).value = tbdlmj_list[i]
            else :
                if worksheet.cell(row+index,5).value != 0.0000:
                    worksheet.cell(row+index,5).value = tbdlmj_list[i] + float(worksheet.cell(row+index,5).value)
                else :
                    worksheet.cell(row+index,5).value = tbdlmj_list[i]
        elif dlmc_list[i] == '盐碱地' and pjjg_list[i] != ' ':
            if pjjg_list[i] == 'FHB' :
                if worksheet.cell(row+index,11).value != 0.0000 :
                    worksheet.cell(row+index,11).value = tbdlmj_list[i] + float(worksheet.cell(row+index,11).value)
                else :
                    worksheet.cell(row+index,11).value = tbdlmj_list[i]
            else :
                if worksheet.cell(row+index,6).value != 0.0000 :
                    worksheet.cell(row+index,6).value = tbdlmj_list[i] + float(worksheet.cell(row+index,6).value)
                else :
                    worksheet.cell(row+index,6).value = tbdlmj_list[i]
        elif dlmc_list[i] == '沙地' and pjjg_list[i] != ' ':
            if pjjg_list[i] == 'FHB' :
                if worksheet.cell(row+index,12).value != 0.0000 :
                    worksheet.cell(row+index,12).value = tbdlmj_list[i] + float(worksheet.cell(row+index,12).value)
                else :
                    worksheet.cell(row+index,12).value = tbdlmj_list[i]
            else :
                if worksheet.cell(row+index,7).value != 0.0000 :
                    worksheet.cell(row+index,7).value = tbdlmj_list[i] + float(worksheet.cell(row+index,7).value)
                else :
                    worksheet.cell(row+index,7).value = tbdlmj_list[i]
        elif dlmc_list[i] == '裸土地'and pjjg_list[i] != ' ':
            if pjjg_list[i] == 'FHB' :
                if worksheet.cell(row+index,13).value != 0.0000 :
                    worksheet.cell(row+index,13).value = tbdlmj_list[i] +float(worksheet.cell(row+index,13).value)
                else :
                    worksheet.cell(row+index,13).value = tbdlmj_list[i]
            else :
                if worksheet.cell(row+index,8).value != 0.0000 :
                    worksheet.cell(row+index,8).value = tbdlmj_list[i] + float(worksheet.cell(row+index,8).value)
                else :
                    worksheet.cell(row+index,8).value = tbdlmj_list[i]
        else:
            if worksheet.cell(row+index,14).value != 0.0000 :
                 worksheet.cell(row+index,14).value = tbdlmj_list[i]+worksheet.cell(row+index,14).value
            else :
                worksheet.cell(row+index,14).value = tbdlmj_list[i]
        worksheet.cell(row+index,4).value = worksheet.cell(row+index,5).value +worksheet.cell(row+index,6).value+worksheet.cell(row+index,7).value+worksheet.cell(row+index,8).value
        worksheet.cell(row+index,9).value = worksheet.cell(row+index,10).value +worksheet.cell(row+index,11).value+worksheet.cell(row+index,12).value+worksheet.cell(row+index,13).value
        worksheet.cell(row+index,3).value = worksheet.cell(row+index,4).value+worksheet.cell(row+index,9).value+worksheet.cell(row+index,14).value

t3=t4=t5=t6=t7=t8=t9=t10=t11=t12=t13=t14=0
for p in range(total_index+1,row+index+1) :
    t3 = t3 + worksheet.cell(p,3).value
    t4 = t4 + worksheet.cell(p,4).value
    t5 = t5 + worksheet.cell(p,5).value
    t6 = t6 + worksheet.cell(p,6).value
    t7 = t7 + worksheet.cell(p,7).value
    t8 = t8 + worksheet.cell(p,8).value
    t9 = t9 + worksheet.cell(p,9).value
    t10 = t10 + worksheet.cell(p,10).value
    t11 = t11 + worksheet.cell(p,11).value
    t12 = t12 + worksheet.cell(p,12).value
    t13 = t13 + worksheet.cell(p,13).value
    t14 = t14 + worksheet.cell(p,14).value
worksheet.cell(total_index,3).value = t3
worksheet.cell(total_index,4).value = t4
worksheet.cell(total_index,5).value = t5
worksheet.cell(total_index,6).value = t6
worksheet.cell(total_index,7).value = t7
worksheet.cell(total_index,8).value = t8
worksheet.cell(total_index,9).value = t9
worksheet.cell(total_index,10).value = t10
worksheet.cell(total_index,11).value = t11
worksheet.cell(total_index,12).value = t12
worksheet.cell(total_index,13).value = t13
worksheet.cell(total_index,14).value = t14

worksheet.cell(index,1).value = '419001'
worksheet.cell(index,2).value = '济源市'
ta3=ta4=ta5=ta6=ta7=ta8=ta9=ta10=ta11=ta12=ta13=ta14=0 
for q in total_arr :
    ta3 = ta3 + worksheet.cell(q,3).value
    ta4 = ta4 + worksheet.cell(q,4).value
    ta5 = ta5 + worksheet.cell(q,5).value
    ta6 = ta6 + worksheet.cell(q,6).value
    ta7 = ta7 + worksheet.cell(q,7).value
    ta8 = ta8 + worksheet.cell(q,8).value
    ta9 = ta9 + worksheet.cell(q,9).value
    ta10 = ta10 + worksheet.cell(q,10).value
    ta11 = ta11 + worksheet.cell(q,11).value
    ta12 = ta12 + worksheet.cell(q,12).value
    ta13 = ta13 + worksheet.cell(q,13).value
    ta14 = ta14 + worksheet.cell(q,14).value
worksheet.cell(index,3).value = ta3
worksheet.cell(index,4).value = ta4
worksheet.cell(index,5).value = ta5
worksheet.cell(index,6).value = ta6
worksheet.cell(index,7).value = ta7
worksheet.cell(index,8).value = ta8
worksheet.cell(index,9).value = ta9
worksheet.cell(index,10).value = ta10
worksheet.cell(index,11).value = ta11
worksheet.cell(index,12).value = ta12
worksheet.cell(index,13).value = ta13
worksheet.cell(index,14).value = ta14


workbook.save(file_excel + '/' + '济源市耕地后备资源调查评价分类型统计表' + '.xlsx')