import os
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd

from openpyxl import Workbook

root = tk.Tk()
root.withdraw()

messagebox.showinfo('提示','请选择需要整合的Excel表！')
datapath = filedialog.askopenfilename() 
#读取文档
df1 = pd.read_excel(datapath, sheet_name=0)
nameList = df1['权利人'].tolist()
xzqList = df1['XZQMC'].tolist()
bzList = df1['备注'].tolist()
dlmcList = df1['DLMC'].tolist() 
tbmjList = df1['TBMJ'].tolist()
print(tbmjList)
# 数组去重

newXzqList = pd.unique(xzqList).tolist()
newBZList = pd.unique(bzList).tolist()
newNameList = pd.unique(nameList).tolist()
newDLList = pd.unique(dlmcList).tolist()
# 创建一个workbook 
workbook = Workbook()
# 创建一个worksheet
worksheet = workbook.active

file_excel = os.path.dirname(datapath)

worksheet.cell(1, 1).value = '权利人'
worksheet.cell(1, 2).value = 'XZQMC'
worksheet.cell(1, 3).value = '备注'
worksheet.cell(1, 4).value = 'TBMJ'  
for i in range(len(newDLList)):
    worksheet.cell(1, 5+i).value = newDLList[i]
      
nameStr = ""
row = 0
allArea = 0
for i in range(len(nameList)):
    if nameList[i] != nameStr :
        row = row + 1
        nameStr = nameList[i]
        worksheet.cell(1+row,1).value = nameList[i]
        worksheet.cell(1+row,2).value = xzqList[i]
        worksheet.cell(1+row,3).value = bzList[i]
        worksheet.cell(1+row,4).value = tbmjList[i]  
        #预设面积为0
        for j in range(len(newDLList)):
            worksheet.cell(1+row,5+j).value = 0
        worksheet.cell(1+row,5+newDLList.index(dlmcList[i])).value = tbmjList[i]
        allArea = tbmjList[i]
    else :
        if worksheet.cell(1+row,5+newDLList.index(dlmcList[i])).value:
            worksheet.cell(1+row,5+newDLList.index(dlmcList[i])).value = worksheet.cell(1+row,5+newDLList.index(dlmcList[i])).value + tbmjList[i]
        else:
            worksheet.cell(1+row,5+newDLList.index(dlmcList[i])).value = tbmjList[i]
        allArea += tbmjList[i]
        worksheet.cell(1+row,4).value = allArea

workbook.save(file_excel + '/' + '变更' + '.xlsx')