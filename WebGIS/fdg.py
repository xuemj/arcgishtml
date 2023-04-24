import os
from tkinter import filedialog
from docx import Document
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
root = tk.Tk()
root.withdraw()

messagebox.showinfo('提示','请选择文件夹！')

Folderpath = filedialog.askdirectory() #获得选择好的文件夹


# 获取filepath文件夹下的所有的文件
def getfilelist(filepath):
    filelist =  os.listdir(filepath)
    files = []
    for i in range(len(filelist)):
        child = os.path.join('%s/%s'%(filepath, filelist[i]))
        if os.path.isdir(child):
            files.extend(getfilelist(child))
        else:
            files.append(child)
    return files

lists = []
for filepath in getfilelist(Folderpath):
    if filepath.endswith("docx") and not filepath.startswith('~$'):
        lists.append(filepath)

# 创建一个workbook 
workbook = Workbook()
# 创建一个worksheet
worksheet = workbook.active
worksheet.column_dimensions['A'].width = 22
row = 0

for filepath in lists:
    try:
        doc = Document(filepath)

        for table in doc.tables :
            if table.cell(1,2).text == '地块\n编码' :
                for i in range(2,len(table.rows)):
                    if table.cell(i,1).text != '':
                        row = row+1
                        worksheet.cell(row,1).value = table.cell(i,2).text
                        worksheet.cell(row,2).value = table.cell(i,3).text
                    else :
                        break
    except:
        print('问题路径---',filepath )

        continue
workbook.save(Folderpath + '/' + '原承包合同面积' + '.xlsx')
messagebox.showinfo('提示','获取原承包合同面积完成！')


