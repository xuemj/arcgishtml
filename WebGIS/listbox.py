import os
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd

from openpyxl  import load_workbook

root = tk.Tk()
root.withdraw()

messagebox.showinfo('提示','请选择Excel表！')
datapath = filedialog.askopenfilename() 
#读取文档

def zdgx(xxxx):
    if xxxx == '薄层型':
        return 0.4
    elif xxxx == '松散型':
        return 0.4
    elif xxxx == '紧实型':
        return 0.6
    elif xxxx == '平层型':
        return 0.5
    elif xxxx == '上紧下松型':
        return 0.7
    elif xxxx == '上松下紧型':
        return 1.0
    elif xxxx == '海绵型':
        return 0.9
    elif xxxx == '夹层型':
        return 0.5
    else :
        print('zdgx--',xxxx)
        return 0
def swdyx(xxxx):
    if xxxx == '丰富':
        return 1.0
    elif xxxx == '一般':
        return 0.7
    elif xxxx == '不丰富':
        return 0.4
    else :
        print('swdyx--',xxxx)
        return 0
def ntlwh(xxxx):
    if xxxx == '高':
        return 1
    elif xxxx == '中':
        return 0.7
    elif xxxx == '低':
        return 0.4
    else :
        print('ntlwh--',xxxx)
        return 0
def zays(xxxx):
    if xxxx == '盐碱':
        return 0.5
    elif xxxx == '瘠薄':
        return 0.65
    elif xxxx == '酸化':
        return 0.7
    elif xxxx == '渍潜':
        return 0.55
    elif xxxx == '障碍层次':
        return 0.6
    elif xxxx == '无':
        return 1.0
    else :
        print('zays--',xxxx)
        return 0
def ggnl(xxxx):
    if xxxx == '充分满足':
        return 1
    elif xxxx == '满足':
        return 0.7
    elif xxxx == '基本满足':
        return 0.5
    elif xxxx == '不满足':
        return 0.3
    else :
        print('ggnl--',xxxx)
        return 0
def psnl(xxxx):
    if xxxx == '充分满足':
        return 1.0
    elif xxxx == '满足':
        return 0.7
    elif xxxx == '基本满足':
        return 0.5
    elif xxxx == '不满足':
        return 0.3
    else :
        print('psnl--',xxxx)
        return 0
def qjcd(xxxx):
    if xxxx == '清洁':
        return 1.0
    elif xxxx == '尚清洁':
        return 0.7
    elif xxxx == '轻度污染':
        return 0.5
    elif xxxx == '中度污染':
        return 0.3
    elif xxxx == '重度污染':
        return 0.0
    else :
        print('qjcd--',xxxx)
        return 0
def dxbw(xxxx):
    if xxxx == '冲积平原':
        return 1.0
    elif xxxx == '河谷平原':
        return 1.0
    elif xxxx == '河谷阶地':
        return 0.9
    elif xxxx == '洪积平原':
        return 0.85
    elif xxxx == '黄平塬':
        return 0.8
    elif xxxx == '黄土塬':
        return 0.8
    elif xxxx == '黄土台塬':
        return 0.7
    elif xxxx == '河漫滩':
        return 0.7
    elif xxxx == '低台地':
        return 0.7
    elif xxxx == '黄土残塬':
        return 0.65
    elif xxxx == '低丘陵':
        return 0.65
    elif xxxx == '黄土坪':
        return 0.65
    elif xxxx == '高台地':
        return 0.65
    elif xxxx == '黄土墹':
        return 0.65
    elif xxxx == '黄土梁':
        return 0.6
    elif xxxx == '高丘陵':
        return 0.6
    elif xxxx == '低山':
        return 0.5
    elif xxxx == '黄土峁':
        return 0.5
    elif xxxx == '固定沙地':
        return 0.4
    elif xxxx == '风蚀地':
        return 0.4
    elif xxxx == '中山':
        return 0.4
    elif xxxx == '半固定沙地':
        return 0.3
    elif xxxx == '流动沙地':
        return 0.2
    elif xxxx == '高山':
        return 0.2
    elif xxxx == '极高山':
        return 0.2   
    else :
        print('dxbw--',xxxx)
        return 0
def gczd(xxxx):
    if xxxx == '砂土':
        return 0.4
    elif xxxx == '砂壤':
        return 0.6
    elif xxxx == '轻壤':
        return 0.85
    elif xxxx == '中壤':
        return 1.0
    elif xxxx == '重壤':
        return 0.8
    elif xxxx == '黏土':
        return 0.6
    else :
        print('gczd--',xxxx)
        return 0 
def hb(u):
    a = 0.000001
    c = 649.407006
    if u<=649.4:
        return 1
    elif u>=3649.4:
        return 0
    else :
       return  float(1/(1+a*(u-c)*(u-c)))
def ph(u):
    a = 0.225097
    c = 6.685037
    if u <= 0.4 or u >= 13.0:
        return 0
    else :
        return float(1/(1+a*(u-c)*(u-c)))
def yjz(u):
    a = 0.006107
    c = 27.680348
    if u <= 0:
        return 0
    elif u >= 27.7:
        return 1
    else :
        return float(1/(1+a*(u-c)*(u-c)))
def sxj(u):
    a = 0.000026
    c = 293.758384
    if u <= 0:
        return 0
    elif u >= 294:
        return 1
    else :
        return float(1/(1+a*(u-c)*(u-c)))
def yxl(u):
    a = 0.001821
    c = 38.076968
    if u <= 0:
        return 0
    elif u >= 38.1:
        return 1
    else :
        return float(1/(1+a*(u-c)*(u-c)))
def trrz(u):
    a = 13.854674
    c = 1.250789
    if u <= 0.44 or u >= 2.05:
        return 0
    else :
        return float(1/(1+a*(u-c)*(u-c)))
def yxtch(u):
    a = 0.000232
    c = 131.349274
    if u <= 0:
        return 0
    elif u >= 131:
        return 1
    else :
        return float(1/(1+a*(u-c)*(u-c)))
def djhf(u):
    if u >= 0.0000 and u < 0.6741:
        return 8
    elif u >= 0.6741 and u < 0.6819:
        return 7
    elif u >= 0.6819 and u < 0.6969:
        return 6
    elif u >= 0.6969 and u < 0.7136:
        return 5
    elif u >= 0.7136 and u < 0.7326:
        return 4
    elif u >= 0.7326 and u < 0.7749:
        return 3
    elif u >= 0.7749 and u < 0.7904:
        return 2
    elif u >= 0.7904 and u < 1.0000:
        return 1
wb = load_workbook(datapath)
sheet = wb['Sheet1']
i = 0
for row in sheet.rows:
    i+=1
    j = 0
    total = 0
    for cell in row:
        if cell.value == '序号':
            break
        j+=1
        if j == 2:
            print('zdgx--',zdgx(cell.value)*0.0694)
            total += zdgx(cell.value)*0.0694 #质地构型
        elif j == 3:
            print('swdyx--',swdyx(cell.value)*0.0303)
            total += swdyx(cell.value)*0.0303 #生物多样性
        elif j == 4:
            print('ntlwh--',ntlwh(cell.value)*0.0384)
            total += ntlwh(cell.value)*0.0384 #农田林网化
        elif j == 5:
            print('zays--',zays(cell.value)*0.0426)
            total += zays(cell.value)*0.0426 #障碍因素
        elif j == 6:
            print('ggnl--',ggnl(cell.value)*0.1165)
            total += ggnl(cell.value)*0.1165 #灌溉能力
        elif j == 7:
            print('psnl--',psnl(cell.value)*0.045)
            total += psnl(cell.value)*0.045 #排水能力
        elif j == 8:
            print('qjcd--',qjcd(cell.value)*0.0251)
            total += qjcd(cell.value)*0.0251 #清洁程度
        elif j == 9:
            print('hb--',hb(cell.value)*0.0712)
            total += hb(cell.value)*0.0712 #海拔
        elif j == 10:
            print('dxbw--',dxbw(cell.value)*0.1303)
            total += dxbw(cell.value)*0.1303 #地形部位
        elif j == 11:
            print('ph--',ph(cell.value)*0.0396)
            total += ph(cell.value)*0.0396 #PH
        elif j == 12:
            print('yjz--',yjz(cell.value)*0.0894)
            total += yjz(cell.value)*0.0894 #有机质
        elif j == 13:
            print('sxj--',sxj(cell.value)*0.0556)
            total += sxj(cell.value)*0.0556 #速效钾
        elif j == 14:
            print('yxl--',yxl(cell.value)*0.0626)
            total += yxl(cell.value)*0.0626 #有效磷
        elif j == 15:
            print('trrz--',trrz(cell.value)*0.044)
            total += trrz(cell.value)*0.044 #土壤容重
        elif j == 16:
            print('yxtch--',yxtch(cell.value)*0.061)
            total += yxtch(cell.value)*0.061 #有效土层厚
        elif j == 17:
            print('gczd--',gczd(cell.value)*0.079)
            total += gczd(cell.value)*0.079 #耕层质地
    if total > 0:
        print('total',total)
        sheet.cell(i,18).value = total
        sheet.cell(i,19).value = djhf(total)
wb.save(datapath)
messagebox.showinfo('提示','耕地质量等级计算完成！')